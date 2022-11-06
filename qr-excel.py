from random import randint
import openpyxl,qrcode
from PIL import Image
from mega import Mega
import os

try:
    os.makedirs('./Generated QR Codes')
except FileExistsError:
    pass
logo = Image.open('logo.jpg')
file = openpyxl.load_workbook('data.xlsx')
sh = file.active

def designQr(qrData,id,logo):
    basewidth = 70
    wpercent = (basewidth/float(logo.size[0]))
    hsize = int((float(logo.size[1])*float(wpercent)))
    logo = logo.resize((basewidth, hsize), Image.Resampling.LANCZOS)
    QRcode = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_H
    )
    QRcode.add_data(qrData)
    QRimg = QRcode.make_image(
        fill_color="black", back_color="white").convert('RGBA')
    pos = ((QRimg.size[0] - logo.size[0]) // 2,
        (QRimg.size[1] - logo.size[1]) // 2)
    QRimg.paste(logo, pos)
    try:
        os.chdir('Generated QR Codes')
    except:
        pass
    QRimg.save(f'{id}.png')

mega = Mega()
m = mega.login("sitew99844@hempyl.com", "TheSant0x")
folderName='Generated QR Codes'
try:
    m.create_folder(folderName)
except:
    pass
folder = m.find(folderName)
doneCnt=failCnt=0
sh.cell(row = 1, column = 4).value = 'ID'
sh.cell(row = 1, column = 5).value = 'QR Link'
for i in range(2, sh.max_row+1):
    name  = sh.cell(row=i,column=1).value
    phone = sh.cell(row=i,column=2).value
    type = sh.cell(row=i,column=3).value
    if name == None or phone == None or type == None:
        pass
    else:
        id=f"IEEE-{''.join(str(randint(0, 9)) for _ in range(4))}"
        sh.cell(row = i, column = 4).value = id
        qrData=f"Name: {name}\nPhone: {phone}\nID: {id}\nType: {type}"
        designQr(qrData,id,logo)
        idqr = f'{id}.png'
        imgDist=m.upload(idqr, folder[0])
        try:
            imgUrl=m.get_upload_link(imgDist)
            doneCnt+=1
        except:
            print("Couldn't Upload QR image, check your internet connection. you stil have a local copy.")
            failCnt+=1
        sh.cell(row = i ,column=5).value = f'=HYPERLINK("{imgUrl}" ,"QR Link")'
        sh.cell(row = i ,column=5).style = 'Hyperlink'
print(f"Succeed: {doneCnt} QR Codes")
print(f"Failed: {failCnt} QR Codes")
file.save('../data.xlsx')  